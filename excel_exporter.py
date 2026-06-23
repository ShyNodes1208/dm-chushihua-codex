from __future__ import annotations

import re
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def export_query_to_excel(
    cache: Dict[str, Any],
    search: str = "",
    visible_columns: Optional[List[str]] = None,
    env: str = "",
    only_failed: bool = False,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    _fill_sheet(ws, cache, search, visible_columns, env, only_failed)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def export_all_to_excel(
    caches: List[Dict[str, Any]],
    search: str = "",
    env: str = "",
    only_failed: bool = False,
) -> BytesIO:
    """把多个查询导出到同一个工作簿，每个查询一个工作表。"""
    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()
    for cache in caches:
        title = _unique_sheet_title(cache.get("query_name", "query"), used_titles)
        ws = wb.create_sheet(title=title)
        _fill_sheet(ws, cache, search, None, env, only_failed)
    if not wb.sheetnames:
        wb.create_sheet(title="empty")
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _prefilter_cache(cache: Dict[str, Any], env: str, only_failed: bool) -> Dict[str, Any]:
    """按环境与“只看失败”过滤实例行，使导出与界面所见一致。"""
    rows = cache.get("rows", [])
    filtered = rows
    if env:
        filtered = [row for row in filtered if str((row or {}).get("env", "")) == env]
    if only_failed:
        filtered = [row for row in filtered if (row or {}).get("error")]
    if filtered is rows:
        return cache
    new_cache = dict(cache)
    new_cache["rows"] = filtered
    return new_cache


def _fill_sheet(
    ws,
    cache: Dict[str, Any],
    search: str,
    visible_columns: Optional[List[str]],
    env: str = "",
    only_failed: bool = False,
) -> None:
    ws.title = ws.title  # 占位，标题已在创建时设置
    cache = _prefilter_cache(cache, env, only_failed)
    table = build_display_table(cache)
    headers = table["headers"]
    rows = filter_rows(table["rows"], search)
    keep = _keep_indexes(headers, visible_columns)

    ws.append([headers[i] for i in keep])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E9EEF7")

    for row in rows:
        ws.append([row[i] if i < len(row) else "" for i in keep])

    for out_index, _ in enumerate(keep, start=1):
        column = get_column_letter(out_index)
        width = max(len(str(ws.cell(row=r, column=out_index).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[column].width = min(max(width + 2, 12), 42)


def _keep_indexes(headers: List[Any], visible_columns: Optional[List[str]]) -> List[int]:
    if not visible_columns:
        return list(range(len(headers)))
    wanted = {str(name) for name in visible_columns}
    keep = [index for index, name in enumerate(headers) if str(name) in wanted]
    return keep or list(range(len(headers)))


def _unique_sheet_title(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/?*\[\]:]", "_", str(name or "query"))[:31] or "query"
    title = base
    suffix = 1
    while title in used:
        suffix += 1
        tail = f"_{suffix}"
        title = base[: 31 - len(tail)] + tail
    used.add(title)
    return title


def build_display_table(cache: Dict[str, Any]) -> Dict[str, List[List[Any]]]:
    if cache.get("column_labels"):
        return build_meta_table(cache)
    return build_flat_table(cache.get("rows", []))


def build_meta_table(cache: Dict[str, Any]) -> Dict[str, List[List[Any]]]:
    labels = cache.get("column_labels") or []
    headers = ["instance", "env"] + labels
    table_rows = []
    for item in cache.get("rows", []):
        data = list(item.get("data") or [])
        if item.get("error"):
            data = [item.get("error")] + [""] * max(0, len(labels) - 1)
        data = normalize_length(data, len(labels))
        table_rows.append([item.get("instance"), item.get("env", ""), *data])
    return {"headers": headers, "rows": table_rows}


def build_flat_table(rows: List[Dict[str, Any]]) -> Dict[str, List[List[Any]]]:
    base_columns: List[Any] = []
    for item in rows:
        if len(item.get("columns") or []) > len(base_columns):
            base_columns = item["columns"]
    if not base_columns:
        max_values = max((len(data_row) for item in rows for data_row in item.get("data") or []), default=0)
        if max_values:
            base_columns = [f"col_{index}" for index in range(1, max_values + 1)]
    headers = ["instance", "env"] + base_columns + ["error"]
    table_rows = []
    for item in rows:
        if item.get("error"):
            table_rows.append([item.get("instance"), item.get("env", "")] + [""] * len(base_columns) + [item.get("error")])
            continue
        data_rows = item.get("data") or []
        if not data_rows:
            table_rows.append([item.get("instance"), item.get("env", "")] + [""] * len(base_columns) + [""])
            continue
        for data_row in data_rows:
            table_rows.append([item.get("instance"), item.get("env", "")] + normalize_length(list(data_row), len(base_columns)) + [""])
    return {"headers": headers, "rows": table_rows}


def normalize_length(values: List[Any], length: int) -> List[Any]:
    if len(values) < length:
        return values + [""] * (length - len(values))
    return values[:length]


def filter_rows(rows: List[List[Any]], search: str) -> List[List[Any]]:
    keyword = (search or "").strip().lower()
    if not keyword:
        return rows
    return [row for row in rows if keyword in " ".join(str(cell).lower() for cell in row)]


if __name__ == "__main__":
    demo = {"query_name": "demo", "column_labels": ["version"], "rows": [{"instance": "DB-01", "env": "测试", "data": ["8"], "error": None}]}
    print(len(export_query_to_excel(demo).getvalue()))
