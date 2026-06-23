import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def clean_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_host(value) -> str:
    return re.sub(r"\s+", "", clean_cell(value))


def load_instances(excel_path: Path, sheet_name: str | None, start_row: int) -> list[dict[str, str]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    instances = []

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=start_row, min_col=1, max_col=3, values_only=True),
        start=start_row,
    ):
        env, label, host = (clean_cell(row[0]), clean_cell(row[1]), clean_host(row[2]))

        if not env and not label and not host:
            continue
        if not label or not host:
            raise ValueError(f"Excel 第 {row_number} 行缺少资产名称或服务器 IP")

        instances.append({"label": label, "host": host, "env": env})

    return instances


def main() -> None:
    parser = argparse.ArgumentParser(
        description="将 Excel 的 A/B/C 列（环境、资产名称、服务器 IP）转换为 config.json"
    )
    parser.add_argument("excel", type=Path, help="Excel 文件路径")
    parser.add_argument(
        "-c",
        "--config",
        type=Path,
        default=Path("config.json"),
        help="作为模板的配置文件，默认 config.json",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("config.generated.json"),
        help="输出文件，默认 config.generated.json",
    )
    parser.add_argument("--sheet", help="工作表名称，默认使用第一个工作表")
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="数据开始行，默认第 2 行（第 1 行为表头）",
    )
    args = parser.parse_args()

    instances = load_instances(args.excel, args.sheet, args.start_row)
    if not instances:
        raise ValueError("Excel 中没有找到有效数据")

    with args.config.open("r", encoding="utf-8") as file:
        config = json.load(file)

    config["instances"] = instances
    with args.output.open("w", encoding="utf-8") as file:
        json.dump(config, file, ensure_ascii=False, indent=2)
        file.write("\n")

    print(f"已生成 {args.output}，共 {len(instances)} 个实例")


if __name__ == "__main__":
    main()
